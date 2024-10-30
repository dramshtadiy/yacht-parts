import requests
from bs4 import BeautifulSoup
import openpyxl
from tqdm import tqdm

arr =[
    'https://yacht-parts.ru//catalog/other_brands/osculati/osculati_02_601_10_vx3_go_24v_1500w_10mm_/',
    'https://yacht-parts.ru//catalog/other_brands/osculati/osculati_02_602_08_vx3_gd_24v_1500w_8mm_/',
    'https://yacht-parts.ru//catalog/other_brands/osculati/osculati_02_602_10_vx3_gd_24v_1500w_10mm_/',
    'https://yacht-parts.ru//catalog/exterior/yakornye_lebedki/os_02_603_06/',
    'https://yacht-parts.ru//catalog/exterior/yakornye_lebedki/os_02_604_06/',
    'https://yacht-parts.ru//catalog/exterior/yakornye_lebedki/os_02_604_08/',
    'https://yacht-parts.ru//catalog/exterior/yakornye_lebedki/os_02_605_06/',
    'https://yacht-parts.ru//catalog/exterior/yakornye_lebedki/os_02_606_06/',
    'https://yacht-parts.ru//catalog/exterior/yakornye_lebedki/os_02_606_08/',
    'https://yacht-parts.ru//catalog/exterior/yakornye_lebedki/os_02_606_99/',
    'https://yacht-parts.ru//catalog/other_brands/osculati/os_02_609_12/',
    'https://yacht-parts.ru//catalog/exterior/yakornye_lebedki/os_02_610_08/',
    'https://yacht-parts.ru//catalog/exterior/yakornye_lebedki/osculati_02_610_12_lewmar_capstan_c3_12_v_/',
    'https://yacht-parts.ru//catalog/exterior/yakornye_lebedki/osculati_02_610_24_lewmar_capstan_c3_24_v_/',
    'https://yacht-parts.ru//catalog/exterior/yakornye_lebedki/osculati_02_612_12_lewmar_capstan_c4_12_v_/',
    'https://yacht-parts.ru//catalog/exterior/yakornye_lebedki/osculati_02_612_24_lewmar_capstan_c4_24_v_/',
    'https://yacht-parts.ru//catalog/exterior/yakornye_lebedki/osculati_02_614_24_lewmar_capstan_c5_24_v_/',
    'https://yacht-parts.ru//catalog/other_brands/osculati/osculati_02_617_24_pyi_tg_winch_24v_c_w_9m_x_8mm_/',
    'https://yacht-parts.ru//catalog/other_brands/osculati/osculati_02_618_12_elephant_tender_capstan_500_w_12_v_/',
    'https://yacht-parts.ru//catalog/other_brands/osculati/osculati_02_618_24_elephant_tender_capstan_500_w_24_v_/',
    'https://yacht-parts.ru//catalog/other_brands/osculati/os_02_619_24/',
    'https://yacht-parts.ru//catalog/exterior/yakornye_lebedki/osculati_02_629_03_08_clipper_1_3kw_24v_10mm_hi_win_/',
    'https://yacht-parts.ru//catalog/other_brands/osculati/osculati_02_636_01_rope_and_connecting_link_10_mm_/',
    'https://yacht-parts.ru//catalog/other_brands/osculati/osculati_02_636_02_rope_and_connecting_link_12_mm_/',
    'https://yacht-parts.ru//catalog/other_brands/osculati/osculati_02_636_05_rope_and_connecting_link_16_mm_/',
    'https://yacht-parts.ru//catalog/exterior/yakornye_lebedki/osculati_02_700_10_built_in_thermal_switch_70_a_/',
    'https://yacht-parts.ru//catalog/exterior/yakornye_lebedki/osculati_02_700_20_built_in_thermal_switch_100_a_/',
    'https://yacht-parts.ru//catalog/exterior/yakornye_lebedki/osculati_02_700_40_built_in_thermal_switch_150_a_/',
    'https://yacht-parts.ru//catalog/exterior/yakornye_lebedki/osculati_02_700_41_built_in_thermal_switch_200_a_/',
    'https://yacht-parts.ru//catalog/exterior/yakornye_lebedki/osculati_02_700_50_built_in_thermal_switch_50_a_/',
    'https://yacht-parts.ru//catalog/exterior/yakornye_lebedki/osculati_02_701_10_external_thermal_switch_70_a_/',
    'https://yacht-parts.ru//catalog/exterior/yakornye_lebedki/osculati_02_701_20_external_thermal_switch_100_a_/',
    'https://yacht-parts.ru//catalog/exterior/yakornye_lebedki/osculati_02_701_40_external_thermal_switch_150_a_/',
    'https://yacht-parts.ru//catalog/exterior/yakornye_lebedki/osculati_02_701_41_external_thermal_switch_200_a_/',
    'https://yacht-parts.ru//catalog/exterior/yakornye_lebedki/osculati_02_701_50_external_thermal_switch_50_a_/',
    'https://yacht-parts.ru//catalog/exterior/yakornye_lebedki/osculati_02_750_00_rubber_seal_for_series_02_750_xx_/',
    'https://yacht-parts.ru//catalog/other_brands/osculati/osculati_02_750_12_built_in_thermal_switch_120_a_/',
    'https://yacht-parts.ru//catalog/other_brands/osculati/osculati_02_750_15_built_in_thermal_switch_150_a_/',
    'https://yacht-parts.ru//catalog/other_brands/osculati/osculati_02_750_20_built_in_thermal_switch_200_a_/',
    'https://yacht-parts.ru//catalog/exterior/yakornye_lebedki/osculati_02_750_60_built_in_thermal_switch_60_a_/',
    'https://yacht-parts.ru//catalog/exterior/yakornye_lebedki/osculati_02_750_80_built_in_thermal_switch_80_a_/',
    'https://yacht-parts.ru//catalog/exterior/yakornye_lebedki/osculati_02_751_12_external_thermal_switch_120_a_/',
    'https://yacht-parts.ru//catalog/exterior/yakornye_lebedki/osculati_02_751_15_external_thermal_switch_150_a_/',
    'https://yacht-parts.ru//catalog/exterior/yakornye_lebedki/osculati_02_751_20_external_thermal_switch_200_a_/',
    'https://yacht-parts.ru//catalog/exterior/yakornye_lebedki/osculati_02_751_60_external_thermal_switch_60_a_/',
    'https://yacht-parts.ru//catalog/exterior/yakornye_lebedki/osculati_02_751_80_external_thermal_switch_80_a_/',
    'https://yacht-parts.ru//catalog/other_brands/osculati/os_02_752_10/',
    'https://yacht-parts.ru//catalog/other_brands/osculati/os_02_752_15/',
    'https://yacht-parts.ru//catalog/other_brands/osculati/os_02_752_20/',]

domen = 'https://yacht-parts.ru/'

def process_card(arr):
    workbook = openpyxl.Workbook()
    sheet = workbook.active

    headers = ['Наименование', 'Цена', 'Артикул', 'Категория', 'Описание', 'Фото']
    for col_num, header in enumerate(headers, 1):
        sheet.cell(row=1, column=col_num, value=header)

    row_num = 2
    for i, card in enumerate(tqdm(arr)):
        response = requests.get(card)
        response.encoding = 'utf-8'
        soup = BeautifulSoup(response.text, 'html.parser')

        name = soup.find('h1', id='pagetitle')
        if name:
            name = name.text.strip()
        else:
            name = 'Нет названия'

        get_price = soup.find('div', class_='middle_info wrap_md')
        if get_price:
            price_1 = get_price.find('div', class_='price')
            if price_1:
                price = price_1.text.strip()
            else:
                price = 'Под заказ'
        else:
            price = ''

        get_article = soup.find('div', class_='article iblock')
        article = get_article.find('span', class_='value').text.strip() if get_article else ''

        get_category = soup.find('div', class_='breadcrumbs')
        category = get_category.find_all('span', itemprop='itemListElement')[3].text.strip() if get_category and len(get_category.find_all('span', itemprop='itemListElement')) > 3 else ''

        desc = soup.find('div', class_='preview_text').text.strip() if soup.find('div', class_='preview_text') else ''

        photo = ''
        get_photo = soup.find('ul', id='thumbs')
        if get_photo:
            img_tags = get_photo.find_all('img')
            photo_urls = [img['src'] if img['src'].startswith('http') else domen + img['src'] for img in img_tags]
            photo = ', '.join(photo_urls)
        else:
            photo_ = soup.find('div', class_='slides')
            if photo_:
                photo_2 = photo_.find('img', attrs={'src': True})
                if photo_2:
                    photo = f'{domen}{photo_2['src']}'

        data = [name, price, article, category, desc, photo]
        for col_num, value in enumerate(data, 1):
            sheet.cell(row=row_num, column=col_num, value=value)

        row_num += 1


    workbook.save('data.xlsx')

process_card(arr)