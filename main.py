import openpyxl
import requests
from bs4 import BeautifulSoup
import concurrent.futures
from tqdm import tqdm

link = 'https://yacht-parts.ru/catalog/'
response = requests.get(link)
response.encoding = 'utf-8'
soup = BeautifulSoup(response.text, 'html.parser')
domen = 'https://yacht-parts.ru/'

link_list = []
all_page = []
all_links = []

def get_first_link(link_list):
    get_subsections = soup.find_all('td', class_='section_info')
    if get_subsections:
        for i in get_subsections:
            get_sect = i.find_all('li', class_='sect')
            if get_sect:
                for ul in get_sect:
                    links = ul.find_all('a', href=True)
                    for a in links:
                        x = f"{domen}{a['href']}"
                        link_list.append(x)
    return link_list

def process_link(one_link):
    response = requests.get(one_link)
    response.encoding = 'utf-8'
    soup = BeautifulSoup(response.text, 'html.parser')
    get_pagination = soup.find('div', class_='module-pagination')
    if get_pagination:
        get_pag_nums = get_pagination.find('span', class_='nums')
        if get_pag_nums:
            last_link = get_pag_nums.find_all('a', href=True)[-1]
            last_num = int(last_link.text) if last_link else 1
        else:
            last_num = 1
    else:
        last_num = 1
    for i in range(1, last_num + 1):
        new_one_link = f'{one_link}?PAGEN_1={i}'
        all_page.append(new_one_link)

def get_one_req(link_list):
    with concurrent.futures.ThreadPoolExecutor(max_workers=10) as executor:
        executor.map(process_link, link_list)
    return all_page

def get_result(all_page):
    for one_page in all_page:
        response = requests.get(one_page)
        response.encoding = 'utf-8'
        soup = BeautifulSoup(response.text, 'html.parser')
        get_item = soup.find_all('div', class_='list_item_wrapp item_wrap')
        if get_item:
            for i in get_item:
                get_last_link = i.find('div', class_='desc_name')
                if get_last_link:
                    a = get_last_link.find('a', href=True)
                    if a:
                        href = a['href']
                        all_link = f'{domen}{href}'
                        print(all_link)
                        all_links.append(all_link)
    return all_links

def process_card(card):
    response = requests.get(card)
    response.encoding = 'utf-8'
    soup = BeautifulSoup(response.text, 'html.parser')

    name = soup.find('h1', id='pagetitle')
    name = name.text.strip() if name else 'Нет названия'

    get_price = soup.find('div', class_='middle_info wrap_md')
    price = get_price.find('div', class_='price').text.strip() if get_price and get_price.find('div', class_='price') else 'Под заказ'

    get_article = soup.find('div', class_='article iblock')
    article = get_article.find('span', class_='value').text.strip() if get_article else ''

    get_category = soup.find('div', class_='breadcrumbs')
    category = get_category.find_all('span', itemprop='itemListElement')[3].text.strip() if get_category and len(get_category.find_all('span', itemprop='itemListElement')) > 3 else ''

    desc = soup.find('div', class_='preview_text').text.strip() if soup.find('div', class_='preview_text') else ''

    photo = ''
    get_photo = soup.find('ul', id='thumbs')
    if get_photo:
        img_tags = get_photo.find_all('img')
        photo_urls = [img['src'] if img['src'].startswith('http') else domen + img['src'] for img in img_tags]
        photo = ', '.join(photo_urls)
    else:
        photo_ = soup.find('div', class_='slides')
        if photo_:
            photo_2 = photo_.find('img', attrs={'src': True})
            if photo_2:
                photo = f'{domen}{photo_2['src']}'
            else:
                photo = 'Без фото'

    data = [name, price, article, category, desc, photo]
    return data

def get_all_items(all_links):
    data = []
    with concurrent.futures.ThreadPoolExecutor(max_workers=15) as executor:
        results = list(tqdm(executor.map(process_card, all_links), total=len(all_links)))
        data.extend(results)
    return data

def save_to_excel(data):
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    headers = ['Наименование', 'Цена', 'Артикул', 'Категория', 'Описание', 'Фото']
    for col_num, header in enumerate(headers, 1):
        sheet.cell(row=1, column=col_num, value=header)
    for row_num, row_data in enumerate(data, 2):
        for col_num, value in enumerate(row_data, 1):
            sheet.cell(row=row_num, column=col_num, value=value)
    workbook.save('data.xlsx')

if __name__ == '__main__':
    get_first_link(link_list)
    get_one_req(link_list)
    all_links = get_result(all_page)
    data = get_all_items(all_links)
    save_to_excel(data)

